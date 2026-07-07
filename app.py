"""
BA Task Manager - Production Backend
Built for Railway.app / Render.com deployment
"""

from flask import Flask, request, jsonify, send_file, render_template_string
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os
import json
from pathlib import Path
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app)

# Configuration
ENV = os.getenv('FLASK_ENV', 'production')
DATA_DIR = Path(os.getenv('DATA_DIR', './data'))
DATA_DIR.mkdir(exist_ok=True)
EXCEL_FILE = DATA_DIR / 'ba_task_manager.xlsx'

# Status colors
STATUS_COLORS = {
    'Lead Assigned': 'DBEAFE',
    'Working on Lead': 'DCFCE7',
    'Document Submitted': 'FEF3C7',
    'Project Lost': 'FEE2E2',
    'Project Won': 'D1FAE5'
}

TASK_HEADERS = [
    'ID', 'Date', 'Task Name', 'Project Name', 'Client Name',
    'Status', 'Estimated Time (min)', 'Actual Time (min)',
    'Technologies', 'Team Members', 'Revenue', 'Lead Source', 'Notes'
]

PROJECT_HEADERS = [
    'ID', 'Project Name', 'Client Name', 'Status', 'Start Date',
    'Technologies', 'Estimated Revenue', 'Lead Source', 'Notes'
]


def init_excel():
    """Initialize Excel file with headers"""
    if not EXCEL_FILE.exists():
        wb = openpyxl.Workbook()
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        # Tasks Sheet
        ws_tasks = wb.active
        ws_tasks.title = 'Tasks'
        ws_tasks.append(TASK_HEADERS)
        for cell in ws_tasks[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        widths = [8, 12, 28, 20, 20, 18, 15, 15, 30, 25, 12, 15, 30]
        for i, width in enumerate(widths, 1):
            ws_tasks.column_dimensions[get_column_letter(i)].width = width
        ws_tasks.row_dimensions[1].height = 30
        ws_tasks.freeze_panes = 'A2'

        # Projects Sheet
        ws_projects = wb.create_sheet('Projects')
        ws_projects.append(PROJECT_HEADERS)
        for cell in ws_projects[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        widths_proj = [8, 20, 20, 18, 12, 30, 16, 15, 30]
        for i, width in enumerate(widths_proj, 1):
            ws_projects.column_dimensions[get_column_letter(i)].width = width
        ws_projects.row_dimensions[1].height = 30
        ws_projects.freeze_panes = 'A2'

        wb.save(EXCEL_FILE)
        logger.info(f"Created new Excel file at {EXCEL_FILE}")


def get_next_id(sheet_name):
    """Get next ID for a specific sheet"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[sheet_name]
    max_id = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            max_id = max(max_id, int(row[0]))
    return max_id + 1


def format_task_row(task):
    """Format task data for Excel"""
    return [
        task.get('id'), task.get('taskDate'), task.get('taskName'),
        task.get('taskProject'), task.get('taskClient', ''),
        task.get('taskStatus'), task.get('taskEstimated'),
        task.get('taskActual'), '; '.join(task.get('technologies', [])),
        '; '.join(task.get('teamMembers', [])), task.get('taskRevenue', 0),
        task.get('taskLead', ''), task.get('taskNotes', '')
    ]


def format_project_row(project):
    """Format project data for Excel"""
    return [
        project.get('id'), project.get('projName'), project.get('projClient'),
        project.get('projStatus'), project.get('projStart'),
        '; '.join(project.get('technologies', [])), project.get('projRevenue', 0),
        project.get('projLead', ''), project.get('projNotes', '')
    ]


def apply_row_style(ws, row_num, status_col):
    """Apply status color to a row"""
    status_cell = ws.cell(row=row_num, column=status_col)
    status = status_cell.value
    if status in STATUS_COLORS:
        fill = PatternFill(start_color=STATUS_COLORS[status], end_color=STATUS_COLORS[status], fill_type='solid')
        status_cell.fill = fill
    for cell in ws[row_num]:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def task_row_to_dict(row):
    """Convert Excel row to task dict"""
    return {
        'id': int(row[0].value),
        'taskDate': str(row[1].value) if row[1].value else '',
        'taskName': row[2].value,
        'taskProject': row[3].value,
        'taskClient': row[4].value or '',
        'taskStatus': row[5].value,
        'taskEstimated': int(row[6].value) if row[6].value else 0,
        'taskActual': int(row[7].value) if row[7].value else 0,
        'technologies': [t.strip() for t in str(row[8].value).split(';') if t.strip()] if row[8].value else [],
        'teamMembers': [m.strip() for m in str(row[9].value).split(';') if m.strip()] if row[9].value else [],
        'taskRevenue': int(row[10].value) if row[10].value else 0,
        'taskLead': row[11].value or '',
        'taskNotes': row[12].value or ''
    }


def project_row_to_dict(row):
    """Convert Excel row to project dict"""
    return {
        'id': int(row[0].value),
        'projName': row[1].value,
        'projClient': row[2].value,
        'projStatus': row[3].value,
        'projStart': str(row[4].value) if row[4].value else '',
        'technologies': [t.strip() for t in str(row[5].value).split(';') if t.strip()] if row[5].value else [],
        'projRevenue': int(row[6].value) if row[6].value else 0,
        'projLead': row[7].value or '',
        'projNotes': row[8].value or ''
    }


# ==================== API ENDPOINTS ====================

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'version': '2.0'})


@app.route('/api/tasks', methods=['GET'])
def get_tasks():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb['Tasks']
        tasks = [task_row_to_dict(row) for row in ws.iter_rows(min_row=2, values_only=False) if row[0].value]
        return jsonify(tasks)
    except Exception as e:
        logger.error(f"Error getting tasks: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/tasks', methods=['POST'])
def add_task():
    try:
        task = request.json
        required = ['taskName', 'taskDate', 'taskProject', 'taskEstimated', 'taskActual', 'taskStatus']
        missing = [f for f in required if not task.get(f) and task.get(f) != 0]
        if missing:
            return jsonify({'error': f'Missing required fields: {", ".join(missing)}'}), 400
        if not task.get('technologies'):
            return jsonify({'error': 'At least one technology is required'}), 400
        if not task.get('teamMembers'):
            return jsonify({'error': 'At least one team member is required'}), 400

        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb['Tasks']
        task['id'] = get_next_id('Tasks')
        row_data = format_task_row(task)
        ws.append(row_data)
        apply_row_style(ws, ws.max_row, status_col=6)
        wb.save(EXCEL_FILE)
        return jsonify({'success': True, 'id': task['id']}), 201
    except Exception as e:
        logger.error(f"Error adding task: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/tasks/<int:task_id>', methods=['PUT'])
def update_task(task_id):
    try:
        updates = request.json
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb['Tasks']

        for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value == task_id:
                updates['id'] = task_id
                row_data = format_task_row(updates)
                for col, value in enumerate(row_data, start=1):
                    ws.cell(row=row_num, column=col, value=value)
                apply_row_style(ws, row_num, status_col=6)
                wb.save(EXCEL_FILE)
                return jsonify({'success': True})

        return jsonify({'error': 'Task not found'}), 404
    except Exception as e:
        logger.error(f"Error updating task: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/tasks/<int:task_id>', methods=['DELETE'])
def delete_task(task_id):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb['Tasks']

        for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value == task_id:
                ws.delete_rows(row_num)
                wb.save(EXCEL_FILE)
                return jsonify({'success': True})

        return jsonify({'error': 'Task not found'}), 404
    except Exception as e:
        logger.error(f"Error deleting task: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/projects', methods=['GET'])
def get_projects():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb['Projects']
        projects = [project_row_to_dict(row) for row in ws.iter_rows(min_row=2, values_only=False) if row[0].value]
        return jsonify(projects)
    except Exception as e:
        logger.error(f"Error getting projects: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/projects', methods=['POST'])
def add_project():
    try:
        project = request.json
        required = ['projName', 'projClient', 'projStatus', 'projStart']
        missing = [f for f in required if not project.get(f)]
        if missing:
            return jsonify({'error': f'Missing required fields: {", ".join(missing)}'}), 400
        if not project.get('technologies'):
            return jsonify({'error': 'At least one technology is required'}), 400

        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb['Projects']
        project['id'] = get_next_id('Projects')
        row_data = format_project_row(project)
        ws.append(row_data)
        apply_row_style(ws, ws.max_row, status_col=4)
        wb.save(EXCEL_FILE)
        return jsonify({'success': True, 'id': project['id']}), 201
    except Exception as e:
        logger.error(f"Error adding project: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/projects/<int:project_id>', methods=['PUT'])
def update_project(project_id):
    try:
        updates = request.json
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb['Projects']

        for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value == project_id:
                updates['id'] = project_id
                row_data = format_project_row(updates)
                for col, value in enumerate(row_data, start=1):
                    ws.cell(row=row_num, column=col, value=value)
                apply_row_style(ws, row_num, status_col=4)
                wb.save(EXCEL_FILE)
                return jsonify({'success': True})

        return jsonify({'error': 'Project not found'}), 404
    except Exception as e:
        logger.error(f"Error updating project: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/projects/<int:project_id>', methods=['DELETE'])
def delete_project(project_id):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb['Projects']

        for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value == project_id:
                ws.delete_rows(row_num)
                wb.save(EXCEL_FILE)
                return jsonify({'success': True})

        return jsonify({'error': 'Project not found'}), 404
    except Exception as e:
        logger.error(f"Error deleting project: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/export', methods=['GET'])
def export_excel():
    try:
        return send_file(EXCEL_FILE, as_attachment=True,
                          download_name=f'ba_task_manager_{datetime.now().strftime("%Y%m%d")}.xlsx')
    except Exception as e:
        logger.error(f"Error exporting Excel: {e}")
        return jsonify({'error': str(e)}), 500


def get_date_range(report_type, report_date_str):
    """Get date range for report"""
    ref_date = datetime.strptime(report_date_str, '%Y-%m-%d').date()

    if report_type == 'daily':
        return ref_date, ref_date
    elif report_type == 'weekly':
        start = ref_date - timedelta(days=(ref_date.weekday() + 1) % 7)
        end = start + timedelta(days=6)
        return start, end
    elif report_type == 'monthly':
        start = ref_date.replace(day=1)
        if start.month == 12:
            next_month = start.replace(year=start.year + 1, month=1)
        else:
            next_month = start.replace(month=start.month + 1)
        end = next_month - timedelta(days=1)
        return start, end
    else:
        return ref_date, ref_date


@app.route('/api/report', methods=['POST'])
def generate_report():
    try:
        data = request.json
        report_type = data.get('type', 'daily')
        report_date = data.get('date')

        start_date, end_date = get_date_range(report_type, report_date)

        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws_tasks = wb['Tasks']

        tasks = []
        for row in ws_tasks.iter_rows(min_row=2, values_only=False):
            if not row[0].value:
                continue
            task_date_str = str(row[1].value)
            try:
                task_date = datetime.strptime(task_date_str, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                continue

            if start_date <= task_date <= end_date:
                tasks.append({
                    'date': task_date_str,
                    'name': row[2].value,
                    'project': row[3].value,
                    'status': row[5].value,
                    'est_time': int(row[6].value) if row[6].value else 0,
                    'actual_time': int(row[7].value) if row[7].value else 0,
                    'revenue': int(row[10].value) if row[10].value else 0,
                    'tech': [t.strip() for t in str(row[8].value).split(';') if t.strip()] if row[8].value else []
                })

        if report_type == 'daily':
            date_label = start_date.strftime('%d %b %Y')
        else:
            date_label = f"{start_date.strftime('%d %b %Y')} - {end_date.strftime('%d %b %Y')}"

        report_text = f"\n{'='*60}\nBA MANAGER REPORT - {report_type.upper()}\n{'='*60}\n"
        report_text += f"Period: {date_label}\nGenerated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"

        report_text += f"SUMMARY\n{'-'*60}\n"
        report_text += f"Total Tasks: {len(tasks)}\n"

        status_count = {}
        for task in tasks:
            status_count[task['status']] = status_count.get(task['status'], 0) + 1
        for status, count in status_count.items():
            report_text += f"{status}: {count}\n"

        total_est = sum(t['est_time'] for t in tasks)
        total_actual = sum(t['actual_time'] for t in tasks)

        report_text += f"\nTIME TRACKING\n{'-'*60}\n"
        report_text += f"Estimated: {total_est} min ({total_est/60:.1f}h)\n"
        report_text += f"Actual: {total_actual} min ({total_actual/60:.1f}h)\n"
        report_text += f"Variance: {total_actual - total_est:+d} min\n"
        if total_actual > 0:
            report_text += f"Efficiency: {(total_est/total_actual)*100:.0f}%\n"

        total_rev = sum(t['revenue'] for t in tasks)
        if total_rev > 0:
            report_text += f"\nREVENUE\n{'-'*60}\nTotal: Rs.{total_rev:,}\n"

        return jsonify({'report': report_text})
    except Exception as e:
        logger.error(f"Error generating report: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/')
def index():
    """Serve the web interface"""
    return render_template_string(get_html_template())


def get_html_template():
    """Return the HTML template"""
    with open(Path(__file__).parent / 'templates' / 'index.html', 'r') as f:
        return f.read()


# ==================== INITIALIZATION ====================

if __name__ == '__main__':
    init_excel()
    port = int(os.getenv('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
